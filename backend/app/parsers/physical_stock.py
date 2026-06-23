from __future__ import annotations

import csv
import io
import unicodedata
from pathlib import Path
from typing import Iterable


MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_DATA_ROWS = 500
SKU_HEADERS = {"sku", "codigo", "codigo_producto"}
QUANTITY_HEADERS = {
    "stock_fisico",
    "fisico",
    "unidades",
    "cantidad",
    "conteo_fisico",
}


class StockFileError(Exception):
    pass


def parse_physical_stock_file(filename: str, content: bytes) -> list[dict[str, object]]:
    if not content:
        raise StockFileError("El archivo esta vacio.")
    if len(content) > MAX_FILE_BYTES:
        raise StockFileError("El archivo supera el limite de 5 MB.")

    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        rows = _csv_rows(content)
    elif suffix == ".xlsx":
        rows = _xlsx_rows(content)
    else:
        raise StockFileError("Formato no permitido. Use CSV o XLSX.")
    return _normalize_rows(rows)


def _csv_rows(content: bytes) -> list[list[object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise StockFileError("El CSV debe estar guardado en formato UTF-8.") from exc

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;")
    except csv.Error:
        dialect = csv.excel
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def _xlsx_rows(content: bytes) -> list[list[object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise StockFileError("Falta instalar openpyxl para leer archivos XLSX.") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True, max_row=MAX_DATA_ROWS + 1)]
    except Exception as exc:
        raise StockFileError("No se pudo leer el archivo XLSX.") from exc


def _normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return "_".join(text.lower().strip().replace("/", " ").split())


def _normalize_rows(rows: Iterable[list[object]]) -> list[dict[str, object]]:
    materialized = [row for row in rows if any(value not in (None, "") for value in row)]
    if len(materialized) < 2:
        raise StockFileError("El archivo debe contener encabezados y al menos una fila.")

    headers = [_normalize_header(value) for value in materialized[0]]
    sku_index = _find_header(headers, SKU_HEADERS)
    quantity_index = _find_header(headers, QUANTITY_HEADERS)
    if sku_index is None or quantity_index is None:
        raise StockFileError("Faltan las columnas SKU y Stock_Fisico.")

    data_rows = materialized[1:]
    if len(data_rows) > MAX_DATA_ROWS:
        raise StockFileError(f"El archivo supera {MAX_DATA_ROWS} filas.")

    normalized: list[dict[str, object]] = []
    for row_number, row in enumerate(data_rows, start=2):
        sku_value = row[sku_index] if sku_index < len(row) else None
        quantity_value = row[quantity_index] if quantity_index < len(row) else None
        if sku_value in (None, "") and quantity_value in (None, ""):
            continue
        sku = str(sku_value or "").upper().strip()
        if not sku:
            raise StockFileError(f"Fila {row_number}: falta el SKU.")
        quantity = _integer_quantity(quantity_value, row_number)
        normalized.append({"sku": sku, "physical_confirmed": quantity, "row_number": row_number})

    if not normalized:
        raise StockFileError("El archivo no contiene filas de stock.")
    return normalized


def _find_header(headers: list[str], accepted: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in accepted:
            return index
    return None


def _integer_quantity(value: object, row_number: int) -> int:
    try:
        number = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError) as exc:
        raise StockFileError(f"Fila {row_number}: el stock fisico no es numerico.") from exc
    if number < 0 or not number.is_integer():
        raise StockFileError(f"Fila {row_number}: el stock debe ser un entero mayor o igual a cero.")
    return int(number)
